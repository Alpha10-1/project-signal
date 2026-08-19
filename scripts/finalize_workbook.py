"""Apply professional formatting to the cleaned Project Signal workbook and
add a cover / README sheet and a data-quality summary sheet."""
import sys
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def wilson_interval(flagged, total, z=1.96):
    """Wilson score confidence interval on the true proportion of records
    with at least one detectable issue. Used to turn 'X% of records were
    flagged' into an honest statement about the uncertainty of that rate,
    rather than presenting the sample rate as if it were exact.
    """
    if total == 0:
        return (0.0, 0.0, 0.0)
    p = flagged / total
    denom = 1 + z ** 2 / total
    centre = p + z ** 2 / (2 * total)
    margin = z * math.sqrt((p * (1 - p) + z ** 2 / (4 * total)) / total)
    lo = max(0.0, (centre - margin) / denom)
    hi = min(1.0, (centre + margin) / denom)
    return (p, lo, hi)

HEADER_FILL = PatternFill(start_color='1F4E5F', end_color='1F4E5F', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', size=16, bold=True, color='1F4E5F')
SUBTITLE_FONT = Font(name='Arial', size=11, bold=True, color='1F4E5F')
WRAP = Alignment(wrap_text=True, vertical='top')
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_data_sheet(ws):
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    max_col = ws.max_column
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = BODY_FONT
    # reasonable column widths
    for i in range(1, max_col + 1):
        letter = get_column_letter(i)
        header_val = ws.cell(row=1, column=i).value or ''
        width = min(max(12, len(str(header_val)) + 4), 32)
        ws.column_dimensions[letter].width = width
    ws.auto_filter.ref = ws.dimensions


def add_cover_sheet(wb, exception_df, sheets_order, run_info=None, total_records=None):
    ws = wb.create_sheet('README', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 100
    r = 1
    ws.cell(row=r, column=1, value='PROJECT SIGNAL — Task 1: Data Detective').font = TITLE_FONT
    r += 2
    ws.cell(row=r, column=1, value='Cleaned dataset, exception log and reconciliation audit trail').font = SUBTITLE_FONT
    r += 2

    if run_info is not None:
        info = dict(zip(run_info['Field'], run_info['Value']))
        ws.cell(row=r, column=1,
                value=f"Cleaning_Version {info.get('Cleaning_Version', '?')}  |  "
                      f"Run {info.get('Run_Timestamp_UTC', '?')} UTC  |  "
                      f"Input SHA256 {str(info.get('Input_SHA256', '?'))[:12]}…").font = BODY_FONT
        r += 1
        if total_records:
            flagged = exception_df.groupby(['Sheet', 'Record_ID']).ngroups
            p, lo, hi = wilson_interval(flagged, total_records)
            ws.cell(row=r, column=1,
                    value=f"Observed flagged-record rate: {p*100:.1f}% "
                          f"(95% confidence interval {lo*100:.1f}%–{hi*100:.1f}% of {total_records} records). "
                          f"This bounds the *detected* issue rate; it is not a guarantee that undetected "
                          f"errors do not remain outside the rule-set below.").font = BODY_FONT
            r += 1
        r += 1
    lines = [
        'How this workbook is organised:',
        '  - "<Sheet>_RAW" tabs: the original data, completely unmodified, one per source sheet.',
        '  - "<Sheet>_Clean" tabs: the same records with added standardised columns (parsed dates, '
        'canonical equipment codes, normalised category labels, unit-converted quantities) and an '
        'Exception_Flags column listing every issue code that applies to that record. Exact duplicate '
        'rows are removed from these tabs but nothing else is deleted or overwritten.',
        '  - "Duplicates_Removed" tab: full audit trail of every row removed as an exact duplicate.',
        '  - "Exception_Log" tab: one row per flagged issue, with a short code, a plain-language '
        'description, and record-level detail. This is the master reconciliation log.',
        '  - "Data_Quality_Summary" tab: counts of each issue type by dataset, for a quick read of '
        'where the data is weakest.',
        '',
        'Principles followed:',
        '  - Raw evidence is preserved. Nothing is silently overwritten - every correction lives '
        'alongside the original value in a new column.',
        '  - Missing values are kept as blank/NaN, never assumed to be zero.',
        '  - Every assumption used to resolve an ambiguous date format, unit or category is documented '
        'in the Cleaning_Assumptions tab of the companion report and in this workbook\'s code comments.',
        '  - Fields holding direct identifiers (names, emails, phone numbers, badge IDs) or proxy '
        'attributes (home zone, contractor group, medical fitness code) are tagged in the _PII_Fields / '
        '_Proxy_Fields / _Sensitive_Fields columns where present, rather than being removed - removal is '
        'a judgement call for the data owner, not something to bake silently into a cleaning script.',
        '',
        f'Exception codes used (see Exception_Log for the "Code" column):',
    ]
    for line in lines:
        ws.cell(row=r, column=1, value=line).alignment = WRAP if line and line[0] == ' ' else Alignment(wrap_text=True)
        ws.cell(row=r, column=1).font = BODY_FONT
        r += 1

    from clean_project_signal import EXCEPTION_CODES
    for code, desc in EXCEPTION_CODES.items():
        ws.cell(row=r, column=1, value=f'  {code}: {desc}').font = BODY_FONT
        r += 1

    r += 1
    total = len(exception_df)
    unique_records = exception_df.groupby(['Sheet', 'Record_ID']).ngroups
    ws.cell(row=r, column=1, value=f'Total flagged issues: {total}  |  Unique records affected: {unique_records} of 405 records across 9 datasets').font = SUBTITLE_FONT
    return ws


def add_summary_sheet(wb, exception_df):
    ws = wb.create_sheet('Data_Quality_Summary', 1)
    pivot = exception_df.pivot_table(index='Sheet', columns='Code', aggfunc='size', fill_value=0)
    pivot['Total'] = pivot.sum(axis=1)
    pivot = pivot.reset_index()
    for j, col in enumerate(pivot.columns, start=1):
        cell = ws.cell(row=1, column=j, value=str(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(pivot.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val).font = BODY_FONT
    total_row = len(pivot) + 2
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Arial', bold=True)
    for j in range(2, len(pivot.columns) + 1):
        col_letter = get_column_letter(j)
        ws.cell(row=total_row, column=j,
                value=f'=SUM({col_letter}2:{col_letter}{total_row-1})').font = Font(name='Arial', bold=True)
    for i in range(1, len(pivot.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = 'A2'


def main(inpath, outpath):
    wb = load_workbook(inpath)
    exception_df = pd.read_excel(inpath, sheet_name='Exception_Log')

    run_info = None
    total_records = None
    if 'Run_Info' in wb.sheetnames:
        run_info = pd.read_excel(inpath, sheet_name='Run_Info')
        try:
            total_records = int(run_info.set_index('Field').loc['Total_Raw_Records', 'Value'])
        except Exception:
            total_records = None

    for ws_name in wb.sheetnames:
        style_data_sheet(wb[ws_name])

    add_summary_sheet(wb, exception_df)
    add_cover_sheet(wb, exception_df, wb.sheetnames, run_info=run_info, total_records=total_records)

    wb.save(outpath)
    print('Saved', outpath)


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
